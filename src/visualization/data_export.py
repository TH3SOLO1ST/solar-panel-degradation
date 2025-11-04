"""
Data Export Module

This module provides comprehensive data export functionality for solar panel
degradation analysis results. It supports multiple formats including CSV,
JSON, Excel, MATLAB, and provides structured data organization.

References:
- Pandas documentation for data export
- Microsoft Excel file format specifications
- MATLAB data file formats
- JSON schema standards for scientific data
- NASA Data Management Standards
"""

import numpy as np
import pandas as pd
import json
from typing import List, Dict, Optional, Tuple, Union, Any
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta
from pathlib import Path
import math

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, Reference
except ImportError:
    openpyxl = None

try:
    import scipy.io
except ImportError:
    scipy = None

from ..degradation.lifetime_model import LifetimeState, LifetimePrediction
from ..degradation.power_calculator import PowerOutput
from ..thermal.thermal_analysis import ThermalState, ThermalCycle
from ..orbital.orbit_propagator import OrbitalState


@dataclass
class ExportConfiguration:
    """Configuration for data export"""
    include_metadata: bool = True
    include_raw_data: bool = True
    include_statistics: bool = True
    date_format: str = "%Y-%m-%d %H:%M:%S"
    precision: int = 6
    timezone: str = "UTC"


@dataclass
class ExportMetadata:
    """Metadata for exported data"""
    export_timestamp: datetime
    software_version: str
    scenario_name: str
    mission_duration_hours: float
    data_types: List[str]
    units: Dict[str, str]
    notes: str


class DataExport:
    """
    Comprehensive data export functionality for solar panel degradation analysis.

    Features:
    - Multiple export formats (CSV, JSON, Excel, MATLAB)
    - Structured data organization
    - Metadata inclusion
    - Statistics and summaries
    - Batch export capabilities
    - Custom formatting options
    """

    def __init__(self, config: Optional[ExportConfiguration] = None):
        """
        Initialize data export

        Args:
            config: Export configuration options
        """
        self.config = config or ExportConfiguration()
        self.supported_formats = ["csv", "json", "excel", "matlab", "xlsx", "hdf5"]

    def export_lifetime_data(self, lifetime_states: List[LifetimeState],
                           filepath: str,
                           format: str = "csv",
                           metadata: Optional[ExportMetadata] = None) -> bool:
        """
        Export lifetime degradation data

        Args:
            lifetime_states: List of lifetime states
            filepath: Output file path
            format: Export format
            metadata: Optional metadata to include

        Returns:
            True if export successful
        """
        if not lifetime_states:
            print("No lifetime data to export")
            return False

        # Prepare data
        df = self._lifetime_states_to_dataframe(lifetime_states)

        # Add metadata if requested
        if self.config.include_metadata:
            if metadata is None:
                metadata = self._create_default_metadata(lifetime_states)

        # Export based on format
        try:
            if format.lower() in ["csv", "txt"]:
                self._export_dataframe_csv(df, filepath, metadata)
            elif format.lower() in ["json"]:
                self._export_dataframe_json(df, filepath, metadata)
            elif format.lower() in ["excel", "xlsx"]:
                self._export_lifetime_excel(df, lifetime_states, filepath, metadata)
            elif format.lower() in ["matlab", "mat"]:
                self._export_lifetime_matlab(lifetime_states, filepath, metadata)
            elif format.lower() in ["hdf5", "h5"]:
                self._export_dataframe_hdf5(df, filepath, metadata)
            else:
                raise ValueError(f"Unsupported export format: {format}")

            print(f"Successfully exported lifetime data to {filepath}")
            return True

        except Exception as e:
            print(f"Export failed: {e}")
            return False

    def export_power_data(self, power_outputs: List[PowerOutput],
                         filepath: str,
                         format: str = "csv",
                         metadata: Optional[ExportMetadata] = None) -> bool:
        """
        Export power output data

        Args:
            power_outputs: List of power output states
            filepath: Output file path
            format: Export format
            metadata: Optional metadata

        Returns:
            True if export successful
        """
        if not power_outputs:
            print("No power data to export")
            return False

        # Prepare data
        df = self._power_outputs_to_dataframe(power_outputs)

        # Add metadata
        if self.config.include_metadata and metadata is None:
            metadata = self._create_power_metadata(power_outputs)

        # Export based on format
        try:
            if format.lower() in ["csv", "txt"]:
                self._export_dataframe_csv(df, filepath, metadata)
            elif format.lower() in ["json"]:
                self._export_dataframe_json(df, filepath, metadata)
            elif format.lower() in ["excel", "xlsx"]:
                self._export_power_excel(df, power_outputs, filepath, metadata)
            elif format.lower() in ["matlab", "mat"]:
                self._export_power_matlab(power_outputs, filepath, metadata)
            else:
                raise ValueError(f"Unsupported export format: {format}")

            print(f"Successfully exported power data to {filepath}")
            return True

        except Exception as e:
            print(f"Export failed: {e}")
            return False

    def export_thermal_data(self, thermal_states: List[ThermalState],
                           thermal_cycles: List[ThermalCycle],
                           filepath: str,
                           format: str = "csv",
                           metadata: Optional[ExportMetadata] = None) -> bool:
        """
        Export thermal analysis data

        Args:
            thermal_states: List of thermal states
            thermal_cycles: List of thermal cycles
            filepath: Output file path
            format: Export format
            metadata: Optional metadata

        Returns:
            True if export successful
        """
        if not thermal_states:
            print("No thermal data to export")
            return False

        try:
            if format.lower() in ["excel", "xlsx"]:
                self._export_thermal_excel(thermal_states, thermal_cycles, filepath, metadata)
            elif format.lower() in ["json"]:
                self._export_thermal_json(thermal_states, thermal_cycles, filepath, metadata)
            else:
                # For CSV, export thermal states and cycles separately
                base_path = Path(filepath)
                thermal_df = self._thermal_states_to_dataframe(thermal_states)
                cycles_df = self._thermal_cycles_to_dataframe(thermal_cycles)

                thermal_path = base_path.with_name(f"{base_path.stem}_thermal{base_path.suffix}")
                cycles_path = base_path.with_name(f"{base_path.stem}_cycles{base_path.suffix}")

                self._export_dataframe_csv(thermal_df, str(thermal_path), metadata)
                self._export_dataframe_csv(cycles_df, str(cycles_path), metadata)

            print(f"Successfully exported thermal data to {filepath}")
            return True

        except Exception as e:
            print(f"Export failed: {e}")
            return False

    def export_complete_analysis(self, lifetime_states: List[LifetimeState],
                               power_outputs: List[PowerOutput],
                               thermal_states: List[ThermalState],
                               thermal_cycles: List[ThermalCycle],
                               orbital_states: List[OrbitalState],
                               base_filepath: str,
                               format: str = "excel") -> Dict[str, bool]:
        """
        Export complete analysis results

        Args:
            lifetime_states: Lifetime degradation data
            power_outputs: Power output data
            thermal_states: Thermal analysis data
            thermal_cycles: Thermal cycles data
            orbital_states: Orbital states
            base_filepath: Base file path (without extension)
            format: Export format

        Returns:
            Dictionary of export results
        """
        results = {}
        base_path = Path(base_filepath)

        # Create metadata
        metadata = self._create_complete_metadata(
            lifetime_states, power_outputs, thermal_states, orbital_states
        )

        if format.lower() in ["excel", "xlsx"]:
            # Single Excel file with multiple sheets
            success = self._export_complete_excel(
                lifetime_states, power_outputs, thermal_states, thermal_cycles,
                orbital_states, f"{base_path}.xlsx", metadata
            )
            results["complete_excel"] = success

        else:
            # Multiple files
            if lifetime_states:
                success = self.export_lifetime_data(
                    lifetime_states, f"{base_path}_lifetime.csv", "csv", metadata
                )
                results["lifetime"] = success

            if power_outputs:
                success = self.export_power_data(
                    power_outputs, f"{base_path}_power.csv", "csv", metadata
                )
                results["power"] = success

            if thermal_states:
                success = self.export_thermal_data(
                    thermal_states, thermal_cycles, f"{base_path}_thermal.csv", "csv", metadata
                )
                results["thermal"] = success

            if orbital_states:
                success = self._export_orbital_data(
                    orbital_states, f"{base_path}_orbital.csv", "csv", metadata
                )
                results["orbital"] = success

        return results

    def _lifetime_states_to_dataframe(self, lifetime_states: List[LifetimeState]) -> pd.DataFrame:
        """Convert lifetime states to pandas DataFrame"""
        data = []
        for state in lifetime_states:
            row = {
                'timestamp': state.time.strftime(self.config.date_format),
                'mission_time_hours': state.mission_time_hours,
                'initial_power_W': state.initial_power_watts,
                'current_power_W': state.current_power_watts,
                'power_degradation_percent': state.power_degradation_percent,
                'efficiency_factor': state.efficiency_factor,
                'radiation_damage_factor': state.radiation_damage,
                'thermal_damage_factor': state.thermal_damage,
                'contamination_damage_factor': state.contamination_damage,
                'aging_damage_factor': state.aging_damage,
                'total_radiation_dose_rads': state.total_radiation_dose_rads,
                'total_thermal_cycles': state.total_thermal_cycles,
                'max_temperature_K': state.max_temperature_K,
                'min_temperature_K': state.min_temperature_K,
                'total_eclipse_hours': state.total_eclipse_hours
            }
            data.append(row)

        return pd.DataFrame(data)

    def _power_outputs_to_dataframe(self, power_outputs: List[PowerOutput]) -> pd.DataFrame:
        """Convert power outputs to pandas DataFrame"""
        data = []
        for output in power_outputs:
            row = {
                'timestamp': output.time.strftime(self.config.date_format),
                'power_W': output.power_watts,
                'power_density_W_m2': output.power_density_wm2,
                'current_A': output.current_amps,
                'voltage_V': output.voltage_volts,
                'efficiency': output.efficiency,
                'efficiency_percent': output.efficiency * 100,
                'irradiance_W_m2': output.irradiance_wm2,
                'temperature_K': output.temperature_K,
                'temperature_C': output.temperature_K - 273.15,
                'degradation_factor': output.degradation_factor,
                'eclipse_fraction': output.eclipse_fraction
            }
            data.append(row)

        return pd.DataFrame(data)

    def _thermal_states_to_dataframe(self, thermal_states: List[ThermalState]) -> pd.DataFrame:
        """Convert thermal states to pandas DataFrame"""
        data = []
        for state in thermal_states:
            row = {
                'timestamp': state.time.strftime(self.config.date_format),
                'temperature_K': state.temperature,
                'temperature_C': state.temperature - 273.15,
                'heat_flux_solar_W_m2': state.heat_flux_solar,
                'heat_flux_albedo_W_m2': state.heat_flux_albedo,
                'heat_flux_earth_ir_W_m2': state.heat_flux_earth_ir,
                'heat_flux_radiated_W_m2': state.heat_flux_radiated,
                'net_heat_flux_W_m2': state.net_heat_flux,
                'eclipse_status': state.eclipse_status,
                'temperature_gradient_K_m': state.temperature_gradient
            }
            data.append(row)

        return pd.DataFrame(data)

    def _thermal_cycles_to_dataframe(self, thermal_cycles: List[ThermalCycle]) -> pd.DataFrame:
        """Convert thermal cycles to pandas DataFrame"""
        data = []
        for cycle in thermal_cycles:
            row = {
                'start_time': cycle.start_time.strftime(self.config.date_format),
                'end_time': cycle.end_time.strftime(self.config.date_format),
                'min_temperature_K': cycle.min_temperature,
                'max_temperature_K': cycle.max_temperature,
                'temperature_swing_K': cycle.temperature_swing,
                'cycle_duration_hours': cycle.cycle_duration,
                'heating_rate_K_hr': cycle.heating_rate,
                'cooling_rate_K_hr': cycle.cooling_rate
            }
            data.append(row)

        return pd.DataFrame(data)

    def _export_dataframe_csv(self, df: pd.DataFrame, filepath: str,
                             metadata: Optional[ExportMetadata] = None):
        """Export DataFrame to CSV format"""
        with open(filepath, 'w') as f:
            if metadata:
                f.write(f"# Solar Panel Degradation Analysis Export\n")
                f.write(f"# Export Time: {metadata.export_timestamp.strftime(self.config.date_format)}\n")
                f.write(f"# Software Version: {metadata.software_version}\n")
                f.write(f"# Scenario: {metadata.scenario_name}\n")
                f.write(f"# Notes: {metadata.notes}\n")
                f.write("#\n")

            df.to_csv(f, index=False)

    def _export_dataframe_json(self, df: pd.DataFrame, filepath: str,
                              metadata: Optional[ExportMetadata] = None):
        """Export DataFrame to JSON format"""
        export_data = {
            'metadata': asdict(metadata) if metadata else {},
            'data': json.loads(df.to_json(orient='records', date_format='iso')),
            'statistics': self._calculate_dataframe_statistics(df)
        }

        with open(filepath, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)

    def _export_lifetime_excel(self, df: pd.DataFrame, lifetime_states: List[LifetimeState],
                              filepath: str, metadata: Optional[ExportMetadata] = None):
        """Export lifetime data to Excel format with multiple sheets"""
        if openpyxl is None:
            raise ImportError("openpyxl required for Excel export")

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        if metadata:
            ws_summary['A1'] = "Solar Panel Degradation Analysis Summary"
            ws_summary['A1'].font = Font(bold=True, size=14)
            ws_summary['A3'] = f"Export Time: {metadata.export_timestamp.strftime(self.config.date_format)}"
            ws_summary['A4'] = f"Software Version: {metadata.software_version}"
            ws_summary['A5'] = f"Scenario: {metadata.scenario_name}"
            ws_summary['A6'] = f"Mission Duration: {metadata.mission_duration_hours:.1f} hours"

        # Add key metrics
        if lifetime_states:
            initial_power = lifetime_states[0].current_power_watts
            final_power = lifetime_states[-1].current_power_watts
            total_degradation = lifetime_states[-1].power_degradation_percent

            ws_summary['A8'] = "Key Metrics:"
            ws_summary['A8'].font = Font(bold=True)
            ws_summary['A9'] = f"Initial Power: {initial_power:.2f} W"
            ws_summary['A10'] = f"Final Power: {final_power:.2f} W"
            ws_summary['A11'] = f"Total Degradation: {total_degradation:.2f}%"

        # Lifetime data sheet
        ws_lifetime = wb.create_sheet("Lifetime Data")
        for r_idx, row in enumerate(df.values, 2):
            for c_idx, value in enumerate(row, 1):
                ws_lifetime.cell(row=r_idx, column=c_idx, value=value)

        # Add headers
        for c_idx, header in enumerate(df.columns, 1):
            cell = ws_lifetime.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True)

        # Mechanisms sheet
        if lifetime_states:
            ws_mechanisms = wb.create_sheet("Degradation Mechanisms")
            mechanisms_data = self._extract_mechanism_data(lifetime_states)
            for r_idx, row in enumerate(mechanisms_data, 2):
                for c_idx, value in enumerate(row, 1):
                    ws_mechanisms.cell(row=r_idx, column=c_idx, value=value)

        wb.save(filepath)

    def _export_complete_excel(self, lifetime_states: List[LifetimeState],
                              power_outputs: List[PowerOutput],
                              thermal_states: List[ThermalState],
                              thermal_cycles: List[ThermalCycle],
                              orbital_states: List[OrbitalState],
                              filepath: str, metadata: Optional[ExportMetadata] = None):
        """Export complete analysis to Excel with multiple sheets"""
        if openpyxl is None:
            raise ImportError("openpyxl required for Excel export")

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Add summary information
        self._create_executive_summary(ws_summary, lifetime_states, power_outputs,
                                      thermal_states, metadata)

        # Add data sheets
        if lifetime_states:
            df_lifetime = self._lifetime_states_to_dataframe(lifetime_states)
            self._add_dataframe_sheet(wb, df_lifetime, "Lifetime Degradation")

        if power_outputs:
            df_power = self._power_outputs_to_dataframe(power_outputs)
            self._add_dataframe_sheet(wb, df_power, "Power Output")

        if thermal_states:
            df_thermal = self._thermal_states_to_dataframe(thermal_states)
            self._add_dataframe_sheet(wb, df_thermal, "Thermal Analysis")

        if thermal_cycles:
            df_cycles = self._thermal_cycles_to_dataframe(thermal_cycles)
            self._add_dataframe_sheet(wb, df_cycles, "Thermal Cycles")

        wb.save(filepath)

    def _export_lifetime_matlab(self, lifetime_states: List[LifetimeState],
                               filepath: str, metadata: Optional[ExportMetadata] = None):
        """Export lifetime data to MATLAB format"""
        if scipy is None:
            raise ImportError("scipy required for MATLAB export")

        # Convert to MATLAB-compatible format
        mat_data = {}

        if lifetime_states:
            times = [(state.time - lifetime_states[0].time).total_seconds() / 3600.0
                     for state in lifetime_states]
            powers = [state.current_power_watts for state in lifetime_states]
            degradations = [state.power_degradation_percent for state in lifetime_states]
            efficiencies = [state.efficiency_factor for state in lifetime_states]
            radiation_dose = [state.total_radiation_dose_rads for state in lifetime_states]
            thermal_cycles = [state.total_thermal_cycles for state in lifetime_states]

            mat_data.update({
                'mission_time_hours': np.array(times),
                'power_W': np.array(powers),
                'degradation_percent': np.array(degradations),
                'efficiency_factor': np.array(efficiencies),
                'radiation_dose_rads': np.array(radiation_dose),
                'thermal_cycles': np.array(thermal_cycles)
            })

        if metadata:
            mat_data['metadata'] = {
                'export_timestamp': metadata.export_timestamp.isoformat(),
                'scenario_name': metadata.scenario_name,
                'mission_duration_hours': metadata.mission_duration_hours
            }

        scipy.io.savemat(filepath, mat_data)

    def _create_default_metadata(self, data: List) -> ExportMetadata:
        """Create default metadata for export"""
        return ExportMetadata(
            export_timestamp=datetime.now(),
            software_version="1.0.0",
            scenario_name="Solar Panel Degradation Analysis",
            mission_duration_hours=len(data) if data else 0,
            data_types=["lifetime_degradation"],
            units={"power": "W", "time": "hours", "temperature": "K", "dose": "rads"},
            notes="Generated by Solar Panel Degradation Model"
        )

    def _calculate_dataframe_statistics(self, df: pd.DataFrame) -> Dict:
        """Calculate statistics for DataFrame"""
        stats = {}
        numeric_columns = df.select_dtypes(include=[np.number]).columns

        for col in numeric_columns:
            stats[col] = {
                'mean': float(df[col].mean()),
                'std': float(df[col].std()),
                'min': float(df[col].min()),
                'max': float(df[col].max()),
                'count': int(df[col].count())
            }

        return stats

    def _add_dataframe_sheet(self, wb: Workbook, df: pd.DataFrame, sheet_name: str):
        """Add DataFrame as new sheet to Excel workbook"""
        ws = wb.create_sheet(sheet_name)

        # Add headers
        for c_idx, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True)

        # Add data
        for r_idx, row in enumerate(df.values, 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    def _create_executive_summary(self, ws, lifetime_states, power_outputs, thermal_states, metadata):
        """Create executive summary sheet"""
        ws['A1'] = "Solar Panel Degradation Analysis"
        ws['A1'].font = Font(bold=True, size=16)

        if metadata:
            ws['A3'] = f"Scenario: {metadata.scenario_name}"
            ws['A4'] = f"Export Date: {metadata.export_timestamp.strftime('%Y-%m-%d')}"
            ws['A5'] = f"Software Version: {metadata.software_version}"

        # Add key results
        row = 7
        ws[f'A{row}'] = "Mission Results Summary"
        ws[f'A{row}'].font = Font(bold=True)

        if lifetime_states:
            initial_state = lifetime_states[0]
            final_state = lifetime_states[-1]

            row += 2
            ws[f'A{row}'] = "Initial Performance:"
            ws[f'B{row}'] = f"{initial_state.current_power_watts:.2f} W"
            row += 1
            ws[f'A{row}'] = "Final Performance:"
            ws[f'B{row}'] = f"{final_state.current_power_watts:.2f} W"
            row += 1
            ws[f'A{row}'] = "Total Degradation:"
            ws[f'B{row}'] = f"{final_state.power_degradation_percent:.2f}%"
            row += 1
            ws[f'A{row}'] = "Mission Duration:"
            ws[f'B{row}'] = f"{final_state.mission_time_hours:.1f} hours"

    def export_analysis_report(self, lifetime_states: List[LifetimeState],
                               power_outputs: List[PowerOutput],
                               thermal_states: List[ThermalState],
                               filepath: str) -> bool:
        """
        Generate comprehensive analysis report

        Args:
            lifetime_states: Lifetime degradation data
            power_outputs: Power output data
            thermal_states: Thermal analysis data
            filepath: Output file path

        Returns:
            True if export successful
        """
        try:
            # Create comprehensive report
            report_data = self._generate_analysis_report(
                lifetime_states, power_outputs, thermal_states
            )

            # Export as JSON
            with open(filepath, 'w') as f:
                json.dump(report_data, f, indent=2, default=str)

            print(f"Successfully exported analysis report to {filepath}")
            return True

        except Exception as e:
            print(f"Report export failed: {e}")
            return False

    def _generate_analysis_report(self, lifetime_states, power_outputs, thermal_states) -> Dict:
        """Generate comprehensive analysis report"""
        report = {
            'executive_summary': {},
            'power_analysis': {},
            'thermal_analysis': {},
            'degradation_mechanisms': {},
            'recommendations': []
        }

        # Executive summary
        if lifetime_states:
            initial_state = lifetime_states[0]
            final_state = lifetime_states[-1]

            report['executive_summary'] = {
                'mission_duration_hours': final_state.mission_time_hours,
                'initial_power_W': initial_state.current_power_watts,
                'final_power_W': final_state.current_power_watts,
                'total_degradation_percent': final_state.power_degradation_percent,
                'average_efficiency': np.mean([state.efficiency_factor for state in lifetime_states]),
                'total_radiation_dose_rads': final_state.total_radiation_dose_rads,
                'total_thermal_cycles': final_state.total_thermal_cycles
            }

        # Power analysis
        if power_outputs:
            powers = [p.power_watts for p in power_outputs]
            efficiencies = [p.efficiency for p in power_outputs]

            report['power_analysis'] = {
                'peak_power_W': max(powers),
                'average_power_W': np.mean(powers),
                'minimum_power_W': min(powers),
                'peak_efficiency': max(efficiencies),
                'average_efficiency': np.mean(efficiencies),
                'total_energy_Wh': np.trapz(powers, dx=1.0)  # Assuming hourly data
            }

        # Thermal analysis
        if thermal_states:
            temperatures = [s.temperature for s in thermal_states]
            eclipse_times = [1 for s in thermal_states if s.eclipse_status]

            report['thermal_analysis'] = {
                'max_temperature_K': max(temperatures),
                'min_temperature_K': min(temperatures),
                'average_temperature_K': np.mean(temperatures),
                'temperature_range_K': max(temperatures) - min(temperatures),
                'eclipse_time_fraction': len(eclipse_times) / len(thermal_states)
            }

        return report